#!/usr/bin/env python3
"""Build operator-oriented XLSX exports for joyus fast casual.

Reads JSON payload from stdin and writes workbook to output_path.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _apply_column_formats(ws, headers: list[str], rows: list[list[Any]], formats: dict[str, str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        number_format = formats.get(header)
        if not number_format:
            continue
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if _is_number(cell.value):
                cell.number_format = number_format


def _build_sheet(wb: Workbook, sheet: dict[str, Any]) -> None:
    name = str(sheet.get("name") or "Sheet")
    headers = [str(h) for h in (sheet.get("headers") or [])]
    rows = sheet.get("rows") or []
    col_widths = sheet.get("col_widths") or []
    formats = sheet.get("formats") or {}

    ws = wb.create_sheet(title=name[:31])
    if headers:
        ws.append(headers)
    for row in rows:
        ws.append(row if isinstance(row, list) else [str(row)])

    ws.freeze_panes = "A2"
    if headers and rows:
        ws.auto_filter.ref = ws.dimensions

    for idx, width in enumerate(col_widths, start=1):
        if isinstance(width, (int, float)):
            ws.column_dimensions[get_column_letter(idx)].width = max(7.0, float(width))

    if isinstance(formats, dict) and headers and rows:
        fmt_map = {str(k): str(v) for k, v in formats.items()}
        _apply_column_formats(ws, headers, rows, fmt_map)


def main() -> int:
    raw = sys.stdin.read()
    if not raw:
        print(json.dumps({"ok": False, "error": "Missing stdin payload"}))
        return 1

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        print(json.dumps({"ok": False, "error": f"Invalid JSON payload: {exc}"}))
        return 1

    output_path = payload.get("output_path")
    sheets = payload.get("sheets") or []

    if not output_path:
        print(json.dumps({"ok": False, "error": "output_path is required"}))
        return 1
    if not isinstance(sheets, list) or not sheets:
        print(json.dumps({"ok": False, "error": "sheets must be a non-empty array"}))
        return 1

    out = Path(output_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in sheets:
        if isinstance(sheet, dict):
            _build_sheet(wb, sheet)

    wb.save(out)
    print(json.dumps({"ok": True, "path": str(out)}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
